import pandas as pd
import os
import xlrd
from openpyxl import load_workbook
import pickle
import math
import xlsxwriter
import numpy as np
import editdistance
np.random.seed(0)

type=2 #0=hierarchy, 1=no hiearchy, 2=all

annotations_loc="/home/willie/workspace/Values_Of_ML/data/raw_annotations"

sheets=[]
names=[]
 
# for annotation_file_name in os.listdir(annotations_loc):
# #     annotation_file=os.path.join(annotations_loc, annotation_file_name)
# #     annotaitons=load_workbook(annotation_file)
# #     u=0
#     name=annotation_file_name.split(' ')[0]
#          
#     print('file', annotation_file_name)
#     annotation_file=os.path.join(annotations_loc, annotation_file_name)
#     annotations=pd.ExcelFile(annotation_file)
#     for sheet_name in annotations.sheet_names:
#         #print('sheet_name', sheet_name)
#         sheet=annotations.parse(sheet_name)
#         if sheet._values.shape[0]>0:
#             paper=sheet._values[0][1]
#             print(paper)
#         sheets.append(sheet._values)
#         names.append(name)
# pickle.dump((names,sheets), open('/home/willie/workspace/Values_Of_ML/data/loaded_sheets.p', 'wb'))

names, sheets=pickle.load(open('/home/willie/workspace/Values_Of_ML/data/loaded_sheets.p', 'rb'))
start_ind=28

assignments=pd.ExcelFile("/home/willie/workspace/Values_Of_ML/data/Assignments.xlsx").parse("Sheet1")._values
paper_names=assignments[3:,3]
unique_paper_names=[]
for pn in paper_names:
    if isinstance(pn, str):
        if not pn.lower() in unique_paper_names:
            unique_paper_names.append(pn.lower())
            
cp_unique_paper_names=unique_paper_names.copy()
# if type==0:
#     value_names=["performance", "generalization", "efficiency", "researcher understanding", "building on past work", "novelty", "robustness"]
#     value_indss=[[31,32,34], [9,10,33], [21,35,36,37,38,39,40,82], [48], [43,44], [7], [11]]
# elif type==1:
#     value_names=["performance", "generalization", "efficiency", "researcher understanding", "building on recent work", "novelty", "robustness"]
#     value_indss=[[31], [9], [35], [48], [44], [7], [11]]
# else:
#     value_names=np.ndarray.tolist(sheets[0][27][7:90])
#     value_indss=[list(range(7, 90))]

value_names=np.ndarray.tolist(sheets[0][27][7:90])
i=0
workbook=xlsxwriter.Workbook(os.path.join(annotations_loc, 'all_sentances.xlsx'))
worksheet=workbook.add_worksheet(value_names[0])
for col in range(len(value_names)):
    worksheet.write(0, 1+col, value_names[col])

print("finished loading")

row_num=1
num_sheets=0
papers=[]
for sheet_ind in range(len(sheets)):
    sheet=sheets[sheet_ind]
    if sheet.shape[0]>0 and sheet.shape[1]>90:
        
        paper=sheet[0][1]
#         b=
#         a=[editdistance.eval(paper, papers[i]) for i in range(papers)]
        
        #print(paper)
        if isinstance(paper, str):
            u=0
            num_sheets+=1
            name=names[sheet_ind]
            sentance_ind=29
#             print(paper.replace("\n", " "))
#             if "Generative Modeling" in paper:
#                 u=0
            if not paper.lower() in unique_paper_names:
                print(paper)
            else:
                unique_paper_names.remove(paper.lower())
            
            if len(papers)>0:
                #min_edit_distance=min([editdistance.eval(sheet[sentance_ind][1], papers[i]) for i in range(len(papers))])
                in_substr=False
                for added_paper in papers:
                    if paper in added_paper or added_paper in paper:
                        in_substr=True
                if in_substr:
                    print(paper)
                else:
                    papers.append(paper)
            
            while(isinstance(sheet[sentance_ind][1], str)):
                worksheet.write(row_num, 0, sheet[sentance_ind][1])
                for col_ind in range(7,90):
                    try:
                        u=sheet[sentance_ind]
                        v=sheet[sentance_ind][col_ind]
                    except Exception as e:
                        u=0
                        print("error!", sentance_ind)
                    worksheet.write(row_num, col_ind-6, np.nan_to_num(sheet[sentance_ind][col_ind]))
                sentance_ind+=1
                row_num+=1
print("included", num_sheets, "sheets")       
workbook.close()

values_list=sheets[0][27]

num_sentaces=0
value_sentances=[[] for i in range(len(value_indss))] #sentance, paper, annotator
for sheet_ind in range(len(sheets)):
    sheet=sheets[sheet_ind]
    if sheet.shape[0]>0:
        paper=sheet[0][1]
        if isinstance(paper, str):
            u=0
            name=names[sheet_ind]
            sentance_ind=29
            while(isinstance(sheet[sentance_ind][1], str)):
                for value_inds in range(len(value_indss)):
                    for value_ind in value_indss[value_inds]:
                        if isinstance(sheet[sentance_ind][value_ind], str):
                            value_sentances[value_inds].append((sheet[sentance_ind][1], sheet[sentance_ind][value_ind], paper, name, values_list[value_ind]))
                sentance_ind+=1
                num_sentaces+=1
print(num_sentaces)

if type:
    workbook=xlsxwriter.Workbook(os.path.join(annotations_loc, 'random_hierarchy_values.xlsx'))
elif type==1:
    workbook=xlsxwriter.Workbook(os.path.join(annotations_loc, 'random_top_values.xlsx'))
else:
    workbook=xlsxwriter.Workbook(os.path.join(annotations_loc, 'random_values.xlsx'))

if type==0 or type==1:  
    for value_ind in range(len(value_sentances)):
        worksheet=workbook.add_worksheet(value_names[value_ind])
        worksheet.write(0, 0, 'Sentence')
        worksheet.write(0, 1, 'Implicit/Explicit')
        worksheet.write(0, 2, 'Paper')
        worksheet.write(0, 3, 'Annotator')
        worksheet.write(0, 4, 'Fine-Grained Value')
        permute=np.random.permutation(len(value_sentances[value_ind]))
        for row_ind in range(len(value_sentances[value_ind])):
            worksheet.write(1+row_ind, 0, value_sentances[value_ind][permute[row_ind]][0])
            worksheet.write(1+row_ind, 1, value_sentances[value_ind][permute[row_ind]][1])
            worksheet.write(1+row_ind, 2, value_sentances[value_ind][permute[row_ind]][2])
            worksheet.write(1+row_ind, 3, value_sentances[value_ind][permute[row_ind]][3])
            worksheet.write(1+row_ind, 4, value_sentances[value_ind][permute[row_ind]][4])
    workbook.close()
else:
    i=0
    worksheet=workbook.add_worksheet(value_names[0])
    worksheet.write(0, 0, 'Sentence')
    worksheet.write(0, 1, 'Implicit/Explicit')
    worksheet.write(0, 2, 'Paper')
    worksheet.write(0, 3, 'Annotator')
    worksheet.write(0, 4, 'Fine-Grained Value')
    for value_ind in range(len(value_sentances)):
        permute=np.random.permutation(len(value_sentances[value_ind]))
        for row_ind in range(len(value_sentances[value_ind])):
            worksheet.write(1+row_ind, 0, value_sentances[value_ind][permute[row_ind]][0])
            worksheet.write(1+row_ind, 1, value_sentances[value_ind][permute[row_ind]][1])
            worksheet.write(1+row_ind, 2, value_sentances[value_ind][permute[row_ind]][2])
            worksheet.write(1+row_ind, 3, value_sentances[value_ind][permute[row_ind]][3])
            worksheet.write(1+row_ind, 4, value_sentances[value_ind][permute[row_ind]][4])
            if i<100:
                #print(value_sentances[value_ind][permute[row_ind]])
                sentance=value_sentances[value_ind][permute[row_ind]][0].replace("&", "and")
                print(f'\midrule \"{sentance}\" & {value_sentances[value_ind][permute[row_ind]][4]} \\\\')
                i+=1
    workbook.close()



u=0
